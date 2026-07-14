import os
import re
import uuid
import threading
import urllib.parse
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, jsonify, send_from_directory
import time
import random

app = Flask(__name__)

# Config folders
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if os.environ.get('VERCEL') or os.environ.get('RENDER') or not os.access(BASE_DIR, os.W_OK):
    UPLOAD_FOLDER = '/tmp/uploads'
    DOWNLOAD_FOLDER = '/tmp/downloads'
else:
    UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
    DOWNLOAD_FOLDER = os.path.join(BASE_DIR, 'downloads')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

# In-memory task tracking
tasks = {}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
}

def parse_symptoms(text):
    if not isinstance(text, str):
        return "", ""
    text = text.strip()
    reason_pattern = re.compile(r'(?i)Reason\s*:\s*')
    action_pattern = re.compile(r'(?i)Action\s*:\s*')
    tools_pattern = re.compile(r'(?i)Tools\s*:\s*')
    end_result_pattern = re.compile(r'(?i)End\s+Result\s*:\s*')
    
    reason_match = reason_pattern.search(text)
    action_match = action_pattern.search(text)
    
    reason_val = ""
    action_val = ""
    
    if reason_match and action_match:
        reason_start = reason_match.end()
        reason_end = action_match.start()
        reason_val = text[reason_start:reason_end].strip()
        action_start = action_match.end()
        next_delims = []
        tools_match = tools_pattern.search(text, action_start)
        if tools_match:
            next_delims.append(tools_match.start())
        end_result_match = end_result_pattern.search(text, action_start)
        if end_result_match:
            next_delims.append(end_result_match.start())
        action_end = min(next_delims) if next_delims else len(text)
        action_val = text[action_start:action_end].strip()
    elif reason_match:
        reason_start = reason_match.end()
        next_delims = []
        tools_match = tools_pattern.search(text, reason_start)
        if tools_match:
            next_delims.append(tools_match.start())
        end_result_match = end_result_pattern.search(text, reason_start)
        if end_result_match:
            next_delims.append(end_result_match.start())
        reason_end = min(next_delims) if next_delims else len(text)
        reason_val = text[reason_start:reason_end].strip()
    elif action_match:
        action_start = action_match.end()
        next_delims = []
        tools_match = tools_pattern.search(text, action_start)
        if tools_match:
            next_delims.append(tools_match.start())
        end_result_match = end_result_pattern.search(text, action_start)
        if end_result_match:
            next_delims.append(end_result_match.start())
        action_end = min(next_delims) if next_delims else len(text)
        action_val = text[action_start:action_end].strip()
    else:
        reason_val = text
        action_val = ""
    return reason_val, action_val

def is_valid_device(device):
    if not isinstance(device, str) or pd.isna(device):
        return False
    d = device.replace('\t', ' ').strip().upper()
    if d in ["", "N/A", "NA", "NAN", "NULL", "NONE"]:
        return False
    if "PREMIUM" in d or "方案" in d or "合約" in d or "EWS" in d:
        return False
    return True

def clean_device_name(device):
    if not is_valid_device(device):
        return ""
    d = device.replace('\t', ' ').strip()
    d = re.sub(r'(?i)^(Apple|Samsung|Asus|OPPO|Gigabyte|Sony|Xiaomi|Realme|Vivo|HTC|Huawei|Google)\s+', '', d)
    words = d.split()
    if not words:
        return ""
    first = words[0]
    if first.lower() in ["iphone", "ipad"]:
        if len(words) > 1 and words[1].isdigit():
            return f"{first} {words[1]}"
        return first
    if len(words) > 1:
        return f"{words[0]} {words[1]}"
    return first

def refine_search_query(device, sub_reason, reason_val):
    if not reason_val or reason_val.strip() == "":
        return sub_reason if pd.notna(sub_reason) else ""
        
    first_line = reason_val.split('\n')[0].strip()
    first_line = re.sub(r'\b\d+[\.、]\s*', ' ', first_line)
    
    heuristics = [
        (r"氣泡|畫面多了氣泡", "螢幕氣泡"),
        (r"重複照片", "刪除重複照片"),
        (r"記憶體空間快滿|照片過多需要清除", "照片過多 刪除重複照片"),
        (r"PDF簽名", "PDF簽名 寄出"),
        (r"實體鍵盤.*螢幕鍵盤", "實體鍵盤 螢幕鍵盤"),
        (r"取消訂閱LINE", "取消訂閱 LINE"),
        (r"LINE.*45度", "LINE視訊 45度"),
        (r"數位萬事通.*開通簡訊", "數位萬事通 簡訊"),
        (r"無法打出@", "無法打出@"),
        (r"小白點|小白球", "小白點 設定"),
        (r"中華電信App.*登入|中華電信.*登入", "中華電信App 登入"),
        (r"漫遊", "申請漫遊"),
        (r"LINE ID.*加好友|LIND ID.*加好友", "LINE ID 加好友"),
        (r"LINE加好友", "LINE加好友"),
        (r"檔案下載.*找到", "檔案下載位置"),
        (r"建立檔案列印", "檔案列印"),
        (r"逗哥", "逗哥配音"),
        (r"WECHAT.*書籤|微信.*書籤", "微信 網頁 Chrome 書籤"),
        (r"關掉畫面中小白點", "關閉小白點"),
        (r"App資料庫", "App資料庫"),
        (r"發一封信給聯合國", "發信給聯合國"),
        (r"註銷裝置.*放心家合約", "放心家 註銷裝置"),
        (r"韌體更新", "韌體更新"),
        (r"重新配對|無法配對", "藍牙 重新配對"),
        (r"熱點", "熱點分享 斷線"),
        (r"備份", "備份 教學")
    ]
    
    for pattern, replacement in heuristics:
        if re.search(pattern, first_line, re.IGNORECASE):
            dev = clean_device_name(device)
            return f"{dev} {replacement}".strip()
            
    clauses = re.split(r'[,，、.。;；!！?？\n\(\)]', first_line)
    ignored_words = {"工程師測試", "測試", "test", "ray測試", "測試用", "test query"}
    
    selected_clause = ""
    for c in clauses:
        c_clean = c.strip()
        if not c_clean:
            continue
        if c_clean.lower() in ignored_words or (len(c_clean) < 3 and len(clauses) > 1):
            continue
        if c_clean.startswith("用戶進線詢問") and len(c_clean) < 10:
            continue
        selected_clause = c_clean
        break
        
    if not selected_clause and clauses:
        selected_clause = clauses[0].strip()
        
    prefixes = [
        r"^用戶(?:進線)?(?:詢問|表示|反應|說明)?:?",
        r"^客戶(?:進線)?(?:詢問|表示|反應|說明)?:?",
        r"^進線(?:詢問|表示|反應|說明)?:?",
        r"^(?:想)?(?:詢問|請教|請示|請問)?:?",
        r"^門市(?:人員)?詢問(?:，|,|：|:| )?",
        r"^測試(?:用)?:?",
        r"^test:?",
        r"^並表示",
        r"^想要",
        r"^想",
        r"^要",
        r"^將協助"
    ]
    
    cleaned = selected_clause
    for p in prefixes:
        cleaned = re.compile(p, re.IGNORECASE).sub('', cleaned).strip()
        
    cleaned = re.compile(r'^(?:問題|步驟|第)?[\d一二三四五六七八九十a-zA-Z]+\s*[\.\:：、\-\_]+\s*').sub('', cleaned).strip()
    cleaned = re.sub(r'[,，、\-\_\:\：\s]+$', '', cleaned)
    
    dev = clean_device_name(device)
    query = f"{dev} {cleaned}".strip()
    
    if len(query) > 28:
        query = query[:28].strip()
        query = re.sub(r'[,，、\-\_\:\：\s]+$', '', query)
    return query

def is_relevant(title, url, symptom_term):
    title_lower = title.lower()
    url_lower = url.lower()
    
    # 1. Exclude ads, search helpers, and general portals
    bad_patterns = [
        "廣告", "help.yahoo", "support.microsoft", "bing.com/aclick", "yahoo.com/kb", 
        "廣告相關", "搜尋說明", "服務說明", "隱私權", "服務條款", "客服", "r.search.yahoo"
    ]
    if any(p in title_lower or p in url_lower for p in bad_patterns):
        return False
        
    if not symptom_term or symptom_term.strip() == "":
        return True
        
    # Brand/Service keyword enforce to prevent false positives (e.g. CHT vs China Airlines)
    brand_keywords = ["中華電信", "神腦", "遠傳", "line", "wechat", "deepseek", "pdf", "facebook", "fb", "google", "apple"]
    for bk in brand_keywords:
        if bk in symptom_term.lower():
            if bk not in title_lower and bk not in url_lower:
                return False
                
    # Check for English keywords of length >= 3 in symptom_term (e.g. DeepSeek, LINE, PDF)
    english_words = re.findall(r'[a-zA-Z0-9]{3,}', symptom_term)
    for ew in english_words:
        if ew.lower() in title_lower or ew.lower() in url_lower:
            return True
            
    # For Chinese keywords, check for 2-character substring matching
    clean_symptom = symptom_term
    for filler in ["如何", "怎麼", "自己", "用戶", "說明", "詢問", "電腦", "手機", "平板"]:
        clean_symptom = clean_symptom.replace(filler, "")
        
    if len(clean_symptom) < 2:
        # Check for single character matches if symptom is short
        return any(c in title for c in symptom_term)
        
    for i in range(len(clean_symptom) - 1):
        sub = clean_symptom[i:i+2]
        if sub in title or sub in url:
            return True
            
    # Fallback: check if any character of the clean symptom matches (excluding common characters and limited to Chinese characters)
    match_chars = [c for c in clean_symptom if '\u4e00' <= c <= '\u9fff' and c not in ["手", "機", "電", "話", "設", "定"]]
    if match_chars and any(c in title for c in match_chars):
        return True
        
    return False

def search_yahoo(query, symptom_term):
    if not query or query.strip() == "":
        return []
    
    import random
    domains = [
        "https://tw.search.yahoo.com/search",
        "https://search.yahoo.com/search",
        "https://hk.search.yahoo.com/search"
    ]
    random.shuffle(domains)
    
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
    ]
    
    for domain in domains:
        url = f"{domain}?p={urllib.parse.quote(query)}"
        for attempt in range(1):
            headers = {
                "User-Agent": random.choice(user_agents),
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
                "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
                "Connection": "keep-alive"
            }
            try:
                r = requests.get(url, headers=headers, timeout=5)
                if r.status_code == 200:
                    soup = BeautifulSoup(r.text, 'html.parser')
                    results = []
                    for a_tag in soup.select('div.compTitle a, h3.title a'):
                        h3 = a_tag.find('h3', class_='title')
                        if not h3:
                            h3 = a_tag.select_one('h3')
                        title = h3.get_text().strip() if h3 else a_tag.get_text().strip()
                        
                        href = a_tag.get('href')
                        if href and href.startswith('http') and 'r.search.yahoo.com' in href:
                            if 'RU=' in href:
                                parts = href.split('RU=')
                                if len(parts) > 1:
                                    real_url = urllib.parse.unquote(parts[1].split('/')[0])
                                    href = real_url
                        
                        if href and href.startswith('http') and not 'search.yahoo.com' in href:
                            if is_relevant(title, href, symptom_term):
                                results.append({'title': title, 'url': href})
                                if len(results) >= 3:
                                    break
                    if results:
                        return results
            except Exception:
                pass
            import time
            time.sleep(0.5)
    return []

def search_bing_mobile(query, symptom_term):
    if not query or query.strip() == "":
        return []
    url = f"https://www.bing.com/search?q={urllib.parse.quote(query)}"
    headers = {
        "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.5 Mobile/15E148 Safari/604.1",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7"
    }
    try:
        r = requests.get(url, headers=headers, timeout=5)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            results = []
            for header_div in soup.find_all('div', class_='b_algoheader'):
                a_tag = header_div.find('a')
                if a_tag and a_tag.get('href'):
                    href = a_tag.get('href')
                    title = a_tag.get_text().strip()
                    h = a_tag.find('h2') or a_tag.find('h3') or a_tag.find('div')
                    if h:
                        title = h.get_text().strip()
                    if href.startswith('http') and 'bing.com' not in href and 'microsoft.com' not in href:
                        if is_relevant(title, href, symptom_term):
                            results.append({'title': title, 'url': href})
                            if len(results) >= 3:
                                break
            return results
    except Exception:
        pass
    return []

def search_hybrid(query, symptom_term):
    res = search_yahoo(query, symptom_term)
    if res:
        return res
    return search_bing_mobile(query, symptom_term)

def get_search_results_with_fallback(sub_reason, question_text, action_text):
    # Space out concurrent requests to avoid rate limits
    time.sleep(random.uniform(0.5, 1.5))
    
    # Try 1: Refined main query based ONLY on user question / problem description
    q1 = refine_search_query("", sub_reason, question_text)
    
    if q1 and q1.strip() != "":
        res = search_hybrid(q1, q1)
        if res:
            return res
            
    # Try 2: Shorter version (split by comma and take first part)
    if q1:
        first_part = re.split(r'[,，]', q1)[0].strip()
        if first_part and first_part != q1 and first_part.strip() != "":
            res = search_hybrid(first_part, q1)
            if res:
                return res
                
    # Try 3: Check if there's any specific term in action_text (問題說明) if q1 failed
    if action_text and action_text.strip() != "":
        q3 = refine_search_query("", sub_reason, action_text)
        if q3 and q3.strip() != "" and q3 != q1:
            res = search_hybrid(q3, q3)
            if res:
                return res
                
    # Soluto cases if search fails can be left empty
    return []

def process_file_task(task_id, file_path):
    tasks[task_id]['status'] = 'processing'
    tasks[task_id]['progress'] = 5
    tasks[task_id]['detail'] = '正在讀取 Excel 檔案...'
    
    try:
        # Load excel using pandas
        df = pd.read_excel(file_path)
        
        # Validate required columns
        required_cols = ['WRAP_UP_CODE', 'ENGAGEMENT_TYPE', 'session_reason', 'session_sub_reason', 'Device', 'symptoms', 'Carrier']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValueError(f"上傳的 Excel 缺少必要的欄位: {', '.join(missing)}")
            
        tasks[task_id]['progress'] = 10
        tasks[task_id]['detail'] = '正在過濾案件資料...'
        
        # Filter
        # 1. WRAP_UP_CODE is RESOLVED
        # 2. ENGAGEMENT_TYPE does not contain SUR
        # 3. session_reason is not '未提供服務'
        w_resolved = df['WRAP_UP_CODE'].astype(str).str.upper() == 'RESOLVED'
        eng_sur = df['ENGAGEMENT_TYPE'].astype(str).str.contains('SUR', case=False, na=False)
        not_unserved = df['session_reason'].astype(str).str.strip() != '未提供服務'
        filtered_df = df[w_resolved & (~eng_sur) & not_unserved].copy()
        
        total_cases = len(filtered_df)
        tasks[task_id]['total_cases'] = total_cases
        
        if total_cases == 0:
            raise ValueError("報表過濾後無符合條件 (RESOLVED 且非 SUR) 的案件。")
            
        # Determine year_month from Date column
        year_month = "202605"
        if 'Date' in filtered_df.columns:
            try:
                first_date = filtered_df['Date'].dropna().iloc[0]
                year_month = pd.to_datetime(first_date).strftime('%Y%m')
            except Exception:
                pass
                
        tasks[task_id]['progress'] = 15
        tasks[task_id]['detail'] = f"共篩選出 {total_cases} 筆案件，準備進行文字解析與搜尋..."
        
        # Split into SOLUTO and 非SOLUTO
        soluto_rows = []
        nonsoluto_rows = []
        
        for idx, row in filtered_df.iterrows():
            reason_val, action_val = parse_symptoms(row.get('symptoms'))
            
            carrier_val = str(row.get('Carrier', '')).upper().strip()
            is_ews_carrier = ('CHT_HOME' in carrier_val) or ('SENAO' in carrier_val)
            
            eng_type = str(row.get('ENGAGEMENT_TYPE', '')).upper()
            is_soluto = ('SOLUTO' in eng_type) and (not is_ews_carrier)
            
            row_data = {
                '主分類': row.get('session_reason'),
                '次分類': row.get('session_sub_reason'),
                '裝置': str(row.get('Device')).replace('\t', ' ').strip() if is_valid_device(row.get('Device')) else "",
                '用戶問題': reason_val,
                '問題說明': action_val,
                'is_soluto': is_soluto
            }
            if is_soluto:
                soluto_rows.append(row_data)
            else:
                nonsoluto_rows.append(row_data)
                
        tasks[task_id]['detail'] = "正在非同步搜尋參考資料中..."
        
        # Only search for SOLUTO cases!
        soluto_results = [[]] * len(soluto_rows)
        processed_count = 0
        total_soluto = len(soluto_rows)
        
        if total_soluto > 0:
            with ThreadPoolExecutor(max_workers=5) as executor:
                future_to_idx = {
                    executor.submit(
                        get_search_results_with_fallback,
                        row['次分類'],
                        row['用戶問題'],
                        row['問題說明']
                    ): i for i, row in enumerate(soluto_rows)
                }
                for future in as_completed(future_to_idx):
                    idx = future_to_idx[future]
                    try:
                        res = future.result()
                        soluto_results[idx] = res
                    except Exception as e:
                        print(f"Soluto Row {idx} search failed: {e}")
                        soluto_results[idx] = []
                    
                    processed_count += 1
                    percent = 15 + int((processed_count / total_soluto) * 70)
                    tasks[task_id]['progress'] = percent
                    tasks[task_id]['detail'] = f"已完成搜尋: {processed_count}/{total_soluto} 筆案件..."
        else:
            tasks[task_id]['progress'] = 85
            
        # Non-SOLUTO cases get empty lists immediately without search!
        nonsoluto_results = [[]] * len(nonsoluto_rows)
        
        # Write to Excel using openpyxl
        tasks[task_id]['detail'] = "正在生成 Excel 檔案並套用格式與樣式..."
        
        wb = openpyxl.Workbook()
        
        # Create worksheets
        ws_soluto = wb.active
        ws_soluto.title = f"{year_month}_SOLUTO"
        
        ws_nonsoluto = wb.create_sheet(title=f"{year_month}_EWS相關")
        
        # Styles
        font_header = Font(name='Noto Sans TC Medium', size=12, bold=True, color='000000')
        font_data = Font(name='Noto Sans TC Medium', size=12, bold=False, color='000000')
        font_link = Font(name='Arial', size=10, bold=False, color='0563C1', underline='single')
        
        fill_col_a = PatternFill(start_color='FFEAD1DC', end_color='FFEAD1DC', fill_type='solid')
        fill_col_b = PatternFill(start_color='FFB7B7B7', end_color='FFB7B7B7', fill_type='solid')
        fill_col_c = PatternFill(start_color='FFD0E0E3', end_color='FFD0E0E3', fill_type='solid')
        fill_col_d = PatternFill(start_color='FFFCE5CD', end_color='FFFCE5CD', fill_type='solid')
        fill_col_e = PatternFill(start_color='FFB6D7A8', end_color='FFB6D7A8', fill_type='solid')
        
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        headers = ['主分類', '次分類', '裝置', '用戶問題', '問題說明', '相關補充']
        
        def write_sheet_data(ws, rows_list, links_list):
            ws.views.sheetView[0].showGridLines = True
            
            # Headers
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = font_header
                cell.alignment = align_center
                cell.border = border_thin
                if col_idx == 1:
                    cell.fill = fill_col_a
                elif col_idx == 2:
                    cell.fill = fill_col_b
                elif col_idx == 3:
                    cell.fill = fill_col_c
                elif col_idx == 4:
                    cell.fill = fill_col_d
                elif col_idx == 5:
                    cell.fill = fill_col_e
            ws.row_dimensions[1].height = 25
            
            # Data
            for r_idx, (row_data, links) in enumerate(zip(rows_list, links_list), 2):
                c1 = ws.cell(row=r_idx, column=1, value=row_data['主分類'])
                c1.font = font_data; c1.alignment = align_center; c1.fill = fill_col_a; c1.border = border_thin
                
                c2 = ws.cell(row=r_idx, column=2, value=row_data['次分類'])
                c2.font = font_data; c2.alignment = align_center; c2.fill = fill_col_b; c2.border = border_thin
                
                c3 = ws.cell(row=r_idx, column=3, value=row_data['裝置'])
                c3.font = font_data; c3.alignment = align_center; c3.fill = fill_col_c; c3.border = border_thin
                
                c4 = ws.cell(row=r_idx, column=4, value=row_data['用戶問題'])
                c4.font = font_data; c4.alignment = align_left; c4.fill = fill_col_d; c4.border = border_thin
                
                c5 = ws.cell(row=r_idx, column=5, value=row_data['問題說明'])
                c5.font = font_data; c5.alignment = align_left; c5.fill = fill_col_e; c5.border = border_thin
                
                for l_idx, link in enumerate(links):
                    col_num = 6 + l_idx
                    cell_l = ws.cell(row=r_idx, column=col_num, value=link['title'])
                    cell_l.font = font_link
                    cell_l.alignment = align_left
                    cell_l.hyperlink = link['url']
                    cell_l.border = border_thin
                    
                # Border placeholders for empty link columns
                for l_idx in range(len(links), 3):
                    col_num = 6 + l_idx
                    cell_empty = ws.cell(row=r_idx, column=col_num, value=None)
                    cell_empty.border = border_thin
                    
                ws.row_dimensions[r_idx].height = 75
                
            # Column widths
            ws.column_dimensions['A'].width = 13
            ws.column_dimensions['B'].width = 23
            ws.column_dimensions['C'].width = 32
            ws.column_dimensions['D'].width = 70
            ws.column_dimensions['E'].width = 104
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 30

        # Write both sheets
        write_sheet_data(ws_soluto, soluto_rows, soluto_results)
        write_sheet_data(ws_nonsoluto, nonsoluto_rows, nonsoluto_results)
        
        # Save output file
        result_filename = f"每月案件問題分享 - {year_month}_{task_id[:8]}.xlsx"
        result_path = os.path.join(DOWNLOAD_FOLDER, result_filename)
        wb.save(result_path)
        
        tasks[task_id]['progress'] = 100
        tasks[task_id]['status'] = 'success'
        tasks[task_id]['result_file'] = result_filename
        tasks[task_id]['detail'] = '報表處理成功，準備下載。'
        
    except Exception as e:
        print(f"Task {task_id} failed: {e}")
        tasks[task_id]['status'] = 'error'
        tasks[task_id]['error_msg'] = str(e)
        
    finally:
        # Clean up uploaded file
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception:
                pass

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '未提供檔案'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '檔名為空'}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '僅能上傳 .xlsx 檔案'}), 400
        
    # Generate unique task id
    task_id = str(uuid.uuid4())
    temp_filename = f"{task_id}.xlsx"
    temp_path = os.path.join(UPLOAD_FOLDER, temp_filename)
    file.save(temp_path)
    
    # Initialize task dict
    tasks[task_id] = {
        'status': 'pending',
        'progress': 0,
        'detail': '等待處理中...',
        'result_file': None,
        'error_msg': None,
        'total_cases': 0
    }
    
    # Start thread
    threading.Thread(target=process_file_task, args=(task_id, temp_path)).start()
    
    return jsonify({'task_id': task_id})

@app.route('/status/<task_id>', methods=['GET'])
def get_status(task_id):
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '找不到該任務'}), 404
    return jsonify(task)

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    return send_from_directory(DOWNLOAD_FOLDER, filename, as_attachment=True)

if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        # CLI Mode: python app.py <input_file_path>
        input_path = sys.argv[1]
        print(f"Starting CLI processing for: {input_path}")
        
        # Ensure folders exist
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
        
        # Mock task dictionary for CLI task
        tasks['cli'] = {
            'status': 'pending',
            'progress': 0,
            'detail': '',
            'result_file': None,
            'error_msg': None,
            'total_cases': 0
        }
        process_file_task('cli', input_path)
        print("Processing finished!")
        status = tasks['cli']['status']
        print(f"Status: {status}")
        if status == 'success':
            print(f"Generated result file: {tasks['cli']['result_file']}")
        else:
            print(f"Error: {tasks['cli']['error_msg']}")
            sys.exit(1)
    else:
        port = int(os.environ.get('PORT', 5000))
        app.run(host='0.0.0.0', port=port, debug=True)
